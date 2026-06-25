#!/usr/bin/env python3
"""
build_tracker_json.py

Reads your private Excel job tracker and builds the public-safe tracker.json
used by maricia.com / analytics.html.

Recommended repo layout:

maricia.com/
  analytics.html
  tracker.json
  tracker_tools/
    build_tracker_json.py
    publish_tracker.py
  private/
    Tracker_v25_Master_Clean_Audit.xlsx   <-- gitignored

Run from repo root or tracker_tools:

  python tracker_tools/build_tracker_json.py
  python tracker_tools/build_tracker_json.py --tracker private/Tracker_v25_Master_Clean_Audit.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl\n\n"
        "Install it with:\n"
        "  python -m pip install openpyxl\n"
    ) from exc


PUBLIC_FIELDS = [
    "company",
    "position",
    "status",
    "statusGroup",
    "priority",
    "dateApplied",
    "monthApplied",
    "daysSinceApplied",
    "source",
    "resumeUsed",
    "lastContact",
    "lastActivity",
    "nextFollowUp",
    "notes",
    "verification",
    "category",
    "location",
    "workType",
    "salaryMin",
    "salaryMax",
    "isActive",
]

HEADER_ALIASES = {
    "company": ["company", "employer", "organization"],
    "position": ["position", "role", "job title", "title"],
    "status": ["status", "application status"],
    "statusGroup": ["status group", "statusgroup", "status category"],
    "priority": ["priority", "rank"],
    "dateApplied": ["date applied", "applied date", "application date"],
    "source": ["source", "job source", "platform"],
    "resumeUsed": ["resume used", "resume", "resume version"],
    "lastContact": ["last contact", "last contacted"],
    "lastActivity": ["last activity", "activity date"],
    "nextFollowUp": ["next follow up", "next follow-up", "follow up", "follow-up"],
    "notes": ["notes", "note", "comments", "comment"],
    "verification": ["verification", "verified"],
    "category": ["category", "job category", "industry"],
    "location": ["location", "job location"],
    "workType": ["work type", "worktype", "remote/hybrid/onsite", "arrangement"],
    "salaryMin": ["salary min", "salary minimum", "min salary"],
    "salaryMax": ["salary max", "salary maximum", "max salary"],
    "isActive": ["is active", "active"],
}

CLOSED_WORDS = [
    "reject",
    "rejected",
    "closed",
    "no longer",
    "not selected",
    "not moving forward",
    "not progress",
    "declined",
    "withdrawn",
]

INTERVIEW_WORDS = [
    "interview",
    "phone screen",
    "screen",
    "recruiter",
    "next step",
    "next steps",
]

RESPONSE_WORDS = [
    "received",
    "under review",
    "interview",
    "phone screen",
    "recruiter",
    "rejected",
    "closed",
    "not selected",
    "candidate",
    "next step",
    "next steps",
]


def repo_root_from_script() -> Path:
    """Assumes this file lives in tracker_tools/ under the website repo."""
    here = Path(__file__).resolve()
    if here.parent.name.lower() in {"tracker_tools", "tools"}:
        return here.parent.parent
    return here.parent


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def normalize_key(header: str) -> Optional[str]:
    h = clean_header(header)
    for key, aliases in HEADER_ALIASES.items():
        if h in aliases:
            return key
    return None


def find_tracker_file(root: Path) -> Path:
    candidates = []
    for folder in [root / "private", root, root / "tracker_tools"]:
        if folder.exists():
            candidates.extend(folder.glob("Tracker*.xlsx"))
            candidates.extend(folder.glob("tracker*.xlsx"))

    candidates = [p for p in candidates if not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(
            "Could not find Tracker*.xlsx. Put your workbook in ./private/ or pass --tracker."
        )
    return max(candidates, key=lambda p: p.stat().st_mtime)


def choose_sheet(wb) -> Any:
    preferred = ["Applications", "Application Tracker", "Tracker", "Master", "Jobs"]
    for name in preferred:
        if name in wb.sheetnames:
            return wb[name]

    # Fall back to the sheet with the most rows and columns.
    return max(wb.worksheets, key=lambda ws: ws.max_row * ws.max_column)


def find_header_row(ws, scan_rows: int = 10) -> int:
    best_row = 1
    best_score = -1

    known_terms = {alias for aliases in HEADER_ALIASES.values() for alias in aliases}
    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        vals = [clean_header(c.value) for c in ws[row_idx]]
        score = sum(1 for v in vals if v in known_terms)
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def to_iso_date(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    # Try common formats.
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    # If the cell already contains a recognizable ISO-ish string, keep the date portion.
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if match:
        return match.group(0)

    return text


def month_label(iso_date: Optional[str]) -> str:
    if not iso_date:
        return "Unknown"
    try:
        dt = datetime.strptime(iso_date[:10], "%Y-%m-%d")
        return dt.strftime("%b %Y")
    except Exception:
        return "Unknown"


def days_since(iso_date: Optional[str], today: date) -> Optional[int]:
    if not iso_date:
        return None
    try:
        dt = datetime.strptime(iso_date[:10], "%Y-%m-%d").date()
        return (today - dt).days
    except Exception:
        return None


def as_bool(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1", "active", "open"}:
        return True
    if text in {"false", "no", "n", "0", "closed", "inactive"}:
        return False
    return None


def as_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    text = re.sub(r"[$,]", "", text)
    match = re.search(r"-?\d+(\.\d+)?", text)
    if not match:
        return None
    num = float(match.group(0))
    return int(num) if num.is_integer() else num


def status_group(status: str) -> str:
    s = status.lower()
    if any(w in s for w in CLOSED_WORDS):
        return "Rejected"
    if any(w in s for w in INTERVIEW_WORDS):
        return "Interview"
    if "under review" in s or "under consideration" in s:
        return "Under Review"
    if "applied" in s or "received" in s or "submitted" in s:
        return "Applied"
    return status.strip() or "Applied"


def is_active_from_status(status: str) -> bool:
    s = status.lower()
    return not any(w in s for w in CLOSED_WORDS)


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_applications(tracker_path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(tracker_path, data_only=True)
    ws = choose_sheet(wb)
    header_row = find_header_row(ws)

    raw_headers = [cell.value for cell in ws[header_row]]
    col_map: Dict[int, str] = {}

    for idx, header in enumerate(raw_headers, start=1):
        key = normalize_key(str(header or ""))
        if key:
            col_map[idx] = key

    if not {"company", "position"}.intersection(set(col_map.values())):
        raise ValueError(
            f"Could not identify tracker columns in sheet '{ws.title}'. "
            "Make sure the sheet has headers like Company, Position, Status, Date Applied."
        )

    rows: List[Dict[str, Any]] = []
    today = date.today()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        raw: Dict[str, Any] = {}
        for idx, cell in enumerate(row, start=1):
            key = col_map.get(idx)
            if key:
                raw[key] = cell.value

        company = safe_text(raw.get("company"))
        position = safe_text(raw.get("position"))

        # Skip blank rows.
        if not company and not position:
            continue

        item: Dict[str, Any] = {}

        for field in PUBLIC_FIELDS:
            value = raw.get(field)

            if field in {"dateApplied", "lastContact", "lastActivity", "nextFollowUp"}:
                item[field] = to_iso_date(value)
            elif field in {"salaryMin", "salaryMax"}:
                item[field] = as_number(value)
            elif field == "isActive":
                item[field] = as_bool(value)
            else:
                item[field] = safe_text(value)

        if not item["status"]:
            item["status"] = "Applied"

        if not item["statusGroup"]:
            item["statusGroup"] = status_group(item["status"])

        if item["isActive"] is None:
            item["isActive"] = is_active_from_status(item["status"])

        item["monthApplied"] = month_label(item.get("dateApplied"))
        item["daysSinceApplied"] = days_since(item.get("dateApplied"), today)

        if not item["lastActivity"]:
            item["lastActivity"] = item.get("lastContact") or item.get("dateApplied")

        if not item["category"]:
            item["category"] = "Uncategorized"

        if not item["location"]:
            item["location"] = "Unknown"

        if not item["workType"]:
            item["workType"] = "Unknown"

        if not item["source"]:
            item["source"] = "Tracker"

        rows.append(item)

    return rows


def count_by(rows: Iterable[Dict[str, Any]], field: str, limit: Optional[int] = None) -> List[Dict[str, Any]]:
    counter = Counter((r.get(field) or "Unknown") for r in rows)
    items = [{"label": k, "count": v} for k, v in counter.most_common()]
    return items[:limit] if limit else items


def sort_latest(rows: List[Dict[str, Any]], limit: int = 10) -> List[Dict[str, Any]]:
    def key(row: Dict[str, Any]) -> str:
        return row.get("lastActivity") or row.get("dateApplied") or "0000-00-00"

    return sorted(rows, key=key, reverse=True)[:limit]


def build_json(applications: List[Dict[str, Any]], tracker_path: Path) -> Dict[str, Any]:
    total = len(applications)
    active = sum(1 for r in applications if r.get("isActive"))
    rejections = sum(1 for r in applications if str(r.get("statusGroup", "")).lower() in {"rejected", "closed"})
    interviews = sum(1 for r in applications if "interview" in str(r.get("statusGroup", "")).lower())

    responses = 0
    for r in applications:
        status = str(r.get("status", "")).lower()
        has_contact = bool(r.get("lastContact"))
        if has_contact or any(word in status for word in RESPONSE_WORDS):
            responses += 1

    pending = max(total - rejections - interviews, 0)
    response_rate = round((responses / total) * 100, 1) if total else 0
    interview_rate = round((interviews / total) * 100, 1) if total else 0

    # Sort status by a friendly order.
    status_counts = count_by(applications, "statusGroup")
    status_order = {
        "Applied": 1,
        "Under Review": 2,
        "Under Consideration": 3,
        "Interview": 4,
        "Rejected": 5,
        "Closed": 6,
    }
    status_counts.sort(key=lambda x: (status_order.get(x["label"], 99), x["label"]))

    tech_stack = [
        {"label": "SQL", "count": sum("sql" in (r.get("notes", "") + " " + r.get("position", "")).lower() for r in applications)},
        {"label": "Power BI", "count": sum("power bi" in (r.get("notes", "") + " " + r.get("position", "")).lower() for r in applications)},
        {"label": "Python", "count": sum("python" in (r.get("notes", "") + " " + r.get("position", "")).lower() for r in applications)},
        {"label": "Data Governance", "count": sum("governance" in (r.get("notes", "") + " " + r.get("position", "")).lower() for r in applications)},
        {"label": "Analytics", "count": sum("analytic" in (r.get("notes", "") + " " + r.get("position", "")).lower() for r in applications)},
        {"label": "ETL / ELT", "count": sum(("etl" in (r.get("notes", "") + " " + r.get("position", "")).lower()) or ("elt" in (r.get("notes", "") + " " + r.get("position", "")).lower()) for r in applications)},
    ]
    tech_stack = [x for x in tech_stack if x["count"] > 0]

    data = {
        "meta": {
            "title": "Job Hunt Command Center",
            "generatedFrom": tracker_path.name,
            "generatedOn": date.today().isoformat(),
            "recordCount": total,
        },
        "summary": {
            "applications": total,
            "active": active,
            "responses": responses,
            "interviews": interviews,
            "rejections": rejections,
            "pending": pending,
            "responseRate": response_rate,
            "interviewRate": interview_rate,
        },
        "currentQuest": {
            "primaryQuest": "Find next data / analytics role",
            "currentRegion": "Remote / West Texas",
            "questStatus": "Actively applying",
            "activeSkills": [
                "SQL",
                "Power BI",
                "Python",
                "Data Governance",
                "ETL / ELT",
                "Business Intelligence",
                "Process Automation",
            ],
        },
        "questLog": [
            {"date": date.today().isoformat(), "event": f"Tracker synced from {tracker_path.name}."},
            {"date": date.today().isoformat(), "event": f"{active} active applications currently open."},
        ],
        "charts": {
            "status": status_counts,
            "category": count_by(applications, "category"),
            "location": count_by(applications, "location"),
            "workType": count_by(applications, "workType"),
            "source": count_by(applications, "source"),
            "monthly": count_by(applications, "monthApplied"),
            "techStack": tech_stack,
            "funnel": [
                {"label": "Applications", "count": total},
                {"label": "Responses", "count": responses},
                {"label": "Interviews", "count": interviews},
                {"label": "Offers", "count": 0},
            ],
        },
        "latestActivity": sort_latest(applications, 10),
        "applications": applications,
    }

    return data


def main() -> int:
    parser = argparse.ArgumentParser(description="Build public tracker.json from private Excel tracker.")
    parser.add_argument("--tracker", help="Path to Tracker Excel workbook.")
    parser.add_argument("--output", help="Path to output tracker.json.")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON. Default is pretty because humans.")
    args = parser.parse_args()

    root = repo_root_from_script()
    tracker_path = Path(args.tracker).resolve() if args.tracker else find_tracker_file(root)
    output_path = Path(args.output).resolve() if args.output else root / "tracker.json"

    applications = read_applications(tracker_path)
    data = build_json(applications, tracker_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print("Tracker JSON generated")
    print(f"  Source : {tracker_path}")
    print(f"  Output : {output_path}")
    print(f"  Rows   : {len(applications)}")
    print(f"  Active : {data['summary']['active']}")
    print(f"  Closed : {data['summary']['rejections']}")
    print(f"  Date   : {data['meta']['generatedOn']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
