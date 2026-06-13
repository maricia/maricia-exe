"""
Convert the Excel job tracker into tracker.json for the website.

Usage:
    python convert_tracker.py Tracker_v15_BulkEnriched.xlsx tracker.json

Requires:
    pip install openpyxl
"""

import json
import sys
from collections import Counter
from datetime import datetime, date
from openpyxl import load_workbook


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def norm_status(status):
    s = str(status or "").strip()
    sl = s.lower()
    if "interview" in sl:
        return "Interview"
    if "reject" in sl or "closed" in sl or "not selected" in sl:
        return "Rejected"
    if "offer" in sl:
        return "Offer"
    if "withdraw" in sl:
        return "Withdrawn"
    if "under" in sl:
        return "Under Consideration"
    if "review" in sl:
        return "Under Review"
    if "action" in sl:
        return "Action Required"
    return s or "Unknown"


def main(input_xlsx, output_json):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb["Tracker_v6_Master"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    records = []
    today = date.today()

    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        company = str(rec.get("Company") or "").strip()
        position = str(rec.get("Position") or "").strip()
        if not company and not position:
            continue

        applied = parse_date(rec.get("Date Applied"))
        last_activity = parse_date(rec.get("Last Activity")) or parse_date(rec.get("Last Contact")) or applied
        status_group = norm_status(rec.get("Status"))
        active = status_group not in ("Rejected", "Withdrawn", "Offer")
        age = (today - applied).days if applied else None

        records.append({
            "company": company,
            "position": position,
            "status": str(rec.get("Status") or "").strip(),
            "statusGroup": status_group,
            "priority": str(rec.get("Priority") or "").strip(),
            "dateApplied": applied.isoformat() if applied else None,
            "monthApplied": applied.strftime("%b %Y") if applied else "Unknown",
            "daysSinceApplied": age,
            "source": str(rec.get("Source") or "Unknown").strip(),
            "lastContact": parse_date(rec.get("Last Contact")).isoformat() if parse_date(rec.get("Last Contact")) else None,
            "lastActivity": last_activity.isoformat() if last_activity else None,
            "notes": str(rec.get("Notes") or "").strip(),
            "category": str(rec.get("Category") or "Uncategorized").strip(),
            "location": str(rec.get("Location") or "Unknown").strip(),
            "workType": str(rec.get("Work Type") or "Unknown").strip(),
            "salaryMin": rec.get("Salary Min"),
            "salaryMax": rec.get("Salary Max"),
            "isActive": active
        })

    status_counts = Counter(r["statusGroup"] for r in records)
    category_counts = Counter(r["category"] for r in records)
    location_counts = Counter(r["location"] for r in records)
    month_counts = Counter(r["monthApplied"] for r in records)

    month_order = sorted(
        [m for m in month_counts if m != "Unknown"],
        key=lambda m: datetime.strptime(m, "%b %Y")
    )
    monthly = [{"label": m, "count": month_counts[m]} for m in month_order]
    if "Unknown" in month_counts:
        monthly.append({"label": "Unknown", "count": month_counts["Unknown"]})

    responses = sum(
        1 for r in records
        if r["lastContact"] or r["statusGroup"] in ("Interview", "Rejected", "Under Review", "Under Consideration")
    )
    interviews = sum(1 for r in records if r["statusGroup"] == "Interview" or "interview" in r["notes"].lower())

    payload = {
        "meta": {
            "title": "Job Hunt Command Center",
            "generatedFrom": input_xlsx,
            "generatedOn": date.today().isoformat(),
            "recordCount": len(records)
        },
        "summary": {
            "applications": len(records),
            "active": sum(1 for r in records if r["isActive"]),
            "responses": responses,
            "interviews": interviews,
            "rejections": status_counts.get("Rejected", 0),
            "pending": len(records) - status_counts.get("Rejected", 0) - interviews
        },
        "charts": {
            "status": [{"label": k, "count": v} for k, v in status_counts.most_common()],
            "category": [{"label": k, "count": v} for k, v in category_counts.most_common()],
            "location": [{"label": k, "count": v} for k, v in location_counts.most_common()],
            "monthly": monthly,
            "funnel": [
                {"label": "Applications", "count": len(records)},
                {"label": "Responses", "count": responses},
                {"label": "Interviews", "count": interviews},
                {"label": "Offers", "count": status_counts.get("Offer", 0)}
            ]
        },
        "latestActivity": sorted(records, key=lambda r: r["lastActivity"] or "0000-00-00", reverse=True)[:10],
        "applications": records
    }

    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


if __name__ == "__main__":
    in_file = sys.argv[1] if len(sys.argv) > 1 else "Tracker_v15_BulkEnriched.xlsx"
    out_file = sys.argv[2] if len(sys.argv) > 2 else "tracker.json"
    main(in_file, out_file)
